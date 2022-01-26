import yfinance as yf
import pandas as pd

from openpyxl import Workbook, load_workbook

file_name = str(input())

wb = load_workbook(file_name)

ws = wb.active

tickers = []

row = 9
while ws['B' + str(row)].value!=None:
    tickers.append(ws['B' + str(row)].value)
    row = row+1

    
total_tickers = len(tickers)    

yf_tick = []



for tick in tickers:
    yf_tick.append(yf.Ticker(tick))



stock_infos = []


for tick in yf_tick:
    stock_infos.append(tick.info)
    

live_price = []


for info in stock_infos:
    live_price.append(info['previousClose'])


beta = []


for info in stock_infos:
    beta.append(info['beta'])
    
end = total_tickers + 9    


for a,b in zip(live_price, range(9,end)):
    ws['C' + str(b)].value = a
    

for a,b in zip(beta, range(9,end)):
    ws['J' + str(b)].value = a
    

wb.save('Fund A.xlsx')
print("All done!!")